import argparse
import os
import re
import collections
import openpyxl
import pickle
import networkx as nx
import matplotlib.pyplot as plt


global args
global processed_data
global all_node_count, node_count_array
global job_id, node_count, tps, nic_count
global wb, ws

# field_name and data_{}nodes_N{}_T{}_t{}
global all_data_dict
global all_ring_dict
global all_tree_dict


def parse_arguments():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input_dir",
                        default=None,
                        type=str,
                        required=True,
                        help="The input data dir. Should contain .log files of format"
                             " nccl_test_${jobid}_${node_count}nodes_T${TPS}_N${NIC_count}.log.")
    parser.add_argument("--output",
                        default="nccl_analysis.xlsx",
                        type=str,
                        help="The output filename.")
    parser.add_argument("--gpus_per_node",
                        default=8,
                        type=int,
                        help="# GPU per node")
    global args
    args = parser.parse_args()
    return args


def comm_first_ranks():
    return range(tps)


def world_rank_to_comm_rank_and_idx(world_rank):
    return world_rank // tps, world_rank % tps


def comm_rank_and_idx_to_world_rank(comm_rank, comm_idx):
    return comm_rank * tps + comm_idx


def read_file(full_file_name):
    world_size = node_count * args.gpus_per_node
    prefix_label_len = 1
    threshold = 10
    while world_size >= threshold:
        prefix_label_len += 1
        threshold *= 10
    logs_by_rank = []
    with open(full_file_name, 'r') as f:
        for line in f.readlines():
            world_rank = int(line[:prefix_label_len])
            real_log = line[prefix_label_len + 2:].splitlines()[0]
            logs_by_rank.append((world_rank, real_log))
    return logs_by_rank


def split_log_by_comm(logs_by_rank):
    comm_logs = []
    for i in range(tps):
        comm_logs.append([])
    for world_rank, line in logs_by_rank:
        comm_rank, comm_idx = world_rank_to_comm_rank_and_idx(world_rank)
        comm_logs[comm_idx].append((comm_rank, world_rank, line))
    return comm_logs


#start_allreduce_pattern = re.compile(r"Starting NCCL AllReduce size=(\d+)\.\.\.")
# should skip first 4 elements.
algo_proto_model_pattern = re.compile(
    r"DGX-(\d+):(\d+):(\d+) \[(\d+)\] enqueue.cc:(\d+) "
    r"NCCL WARN Algo=(Tree|Ring), Protocol=(LL|LL128|Simple), time=([-+]?(\d+(\.\d*)?|\.\d+))")
final_algo_proto_model_pattern = re.compile(
    r"DGX-(\d+):(\d+):(\d+) \[(\d+)\] enqueue.cc:(\d+) "
    r"NCCL WARN Final Algo=(Tree|Ring), Protocol=(LL|LL128|Simple), time=(\d+(\.\d*)?|\.\d+)")
pattern_steps_pattern = re.compile(
    r"DGX-(\d+):(\d+):(\d+) \[(\d+)\] enqueue.cc:(\d+) "
    r"NCCL WARN Rank=0, (pattern|nstepsPerLoop)=(\d+)")
perf_result_pattern = re.compile(
    r"Perf result: NCCLGroup=(\d+), Size=(\d+), AlgoBW= (\d+(\.\d*)?|\.\d+) GB/s,"
    r" BusBW= (\d+(\.\d*)?|\.\d+) GB/s, Time\(us\)=(\d+)")


def filter_and_match_comm_log(rank_logs, pattern, log_idx, log_len, filter_comm_rank=None):
    match_info = None
    while log_idx < log_len and match_info is None:
        # has filter and not match
        if filter_comm_rank is not None and filter_comm_rank != rank_logs[log_idx][0]:
            log_idx += 1
            continue
        match_info = pattern.match(rank_logs[log_idx][2])
        log_idx += 1
    return match_info, log_idx


TestDataRecord = collections.namedtuple('TestDataRecord',
                                        ['tps', 'nic', 'node_count', 'data_size', 'comm_rank',
                                         'preds', 'final_algo', 'final_proto', 'final_pred_time',
                                         'pattern', 'nstepsPerLoop',
                                         'algo_bw', 'bus_bw', 'real_time_us'])


def extract_test_results(rank_logs):
    # rank_logs is (comm_rank, world_rank, line)
    log_len = len(rank_logs)
    log_idx = 0
    records = []
    while log_idx < log_len:
        # find 6 projections
        num_projs = 0
        est_data = []
        while num_projs < 6 and log_idx < log_len:
            match_info, log_idx = filter_and_match_comm_log(rank_logs, algo_proto_model_pattern, log_idx, log_len, 0)
            if match_info is not None:
                num_projs += 1
                _, _, _, _, _, algo, proto, est_time_str, _, _ = match_info.groups()
                est_time = float(est_time_str)
                est_data.append((algo, proto, est_time))
        if num_projs != 6:
            assert num_projs == 0
            break
        match_info, log_idx = filter_and_match_comm_log(rank_logs, final_algo_proto_model_pattern, log_idx, log_len, 0)
        assert match_info is not None
        _, _, _, _, _, final_algo, final_proto, pred_time_str, _ = match_info.groups()
        pred_time = float(pred_time_str)
        match_info, log_idx = filter_and_match_comm_log(rank_logs, pattern_steps_pattern, log_idx, log_len, 0)
        assert match_info is not None
        _, _, _, _, _, type_str, final_pattern_str = match_info.groups()
        assert type_str == 'pattern'
        final_pattern = int(final_pattern_str)
        match_info, log_idx = filter_and_match_comm_log(rank_logs, pattern_steps_pattern, log_idx, log_len, 0)
        assert match_info is not None
        _, _, _, _, _, type_str, final_steps_str = match_info.groups()
        assert type_str == 'nstepsPerLoop'
        final_nstepsPerLoop = int(final_steps_str)
        match_info, log_idx = filter_and_match_comm_log(rank_logs, perf_result_pattern, log_idx, log_len, 0)
        assert match_info is not None
        nccl_group_str, data_size_str, algo_bw_str, _, bus_bw_str, _, time_us_str = match_info.groups()
        nccl_group, data_size, algo_bw, bus_bw, real_time_us = \
            int(nccl_group_str), int(data_size_str), float(algo_bw_str), float(bus_bw_str), int(time_us_str)
        test_data_record = TestDataRecord(tps=tps, nic=nic_count, node_count=node_count,
                                          data_size=data_size, comm_rank=nccl_group,
                                          preds=est_data, final_algo=final_algo,
                                          final_proto=final_proto, final_pred_time=pred_time,
                                          pattern=final_pattern, nstepsPerLoop=final_nstepsPerLoop,
                                          algo_bw=algo_bw, bus_bw=bus_bw, real_time_us=real_time_us)
        records.append(test_data_record)
    if len(records) > 0:
        all_node_count[node_count] = 1 if node_count not in all_node_count.keys() else all_node_count[node_count] + 1
        processed_data[node_count] = [records] if node_count not in processed_data.keys() else processed_data[node_count] + [records]


global tree_patterns
ring_pattern = re.compile(
    r"DGX-(\d+):(\d+):(\d+) \[(\d+)\] NCCL INFO Ring (\d+) : (\d+) -> (\d+) -> (\d+)")


def get_node_name_by_comm_rank_and_idx(crank, cidx):
    wr = comm_rank_and_idx_to_world_rank(crank, cidx)
    #return 'CR%dG%dWR%d\nNODE%dGPU%d' % (crank, cidx, wr, wr // 8, wr % 8)
    return 'Rank%d\nNODE%dGPU%d' % (wr, wr // 8, wr % 8)


def create_tree_and_ring_patterns():
    global tree_patterns
    tree_patterns = [None] * 9
    half_channel_count=1
    while half_channel_count <= 8:
        rstr = r"DGX-(\d+):(\d+):(\d+) \[(\d+)\] NCCL INFO Trees"
        for i in range(half_channel_count * 2):
            rstr += r" \[(\d+)\] ([-+]?\d+)/([-+]?\d+)/([-+]?\d+)->([-+]?\d+)->([-+]?\d+)"
        tree_patterns[half_channel_count] = re.compile(rstr)
        half_channel_count *= 2


def extract_tree_info(comm_idx_logs, comm_idx):
    # rank_logs is (comm_rank, world_rank, line)
    log_len = len(comm_idx_logs)
    g_trees = all_tree_dict['tree_%dnodes_N%d_T%d_t%d' % (node_count, nic_count, tps, comm_idx)]
    tree_channel_count = max(nic_count // tps, 1)
    for log_idx in range(log_len):
        log_str = comm_idx_logs[log_idx]
        comm_rank = log_str[0]
        world_rank = log_str[1]
        match_info = tree_patterns[tree_channel_count].match(log_str[2])
        if match_info is None:
            continue
        mg = match_info.groups()
        _, _, _, _ = mg[:4]
        assert world_rank == comm_rank_and_idx_to_world_rank(comm_rank, comm_idx)
        for ch in range(tree_channel_count * 2):
            channel_idx, bottom0, bottom1, bottom2, current_node, parent_node = mg[4 + ch * 6: 4 + ch * 6 + 6]
            channel_idx, bottom0, bottom1, bottom2, current_node, parent_node =\
                (int(channel_idx), int(bottom0), int(bottom1), int(bottom2), int(current_node), int(parent_node))
            assert channel_idx == ch
            assert current_node == comm_rank
            if parent_node >= 0:
                g_trees[ch].add_edge(get_node_name_by_comm_rank_and_idx(current_node, comm_idx),
                                     get_node_name_by_comm_rank_and_idx(parent_node, comm_idx))


def extract_ring_info(comm_idx_logs, comm_idx):
    # rank_logs is (comm_rank, world_rank, line)
    log_len = len(comm_idx_logs)
    ring_name = 'ring_%dnodes_N%d_T%d_t%d' % (node_count, nic_count, tps, comm_idx)
    g_rings = all_ring_dict[ring_name]
    for log_idx in range(log_len):
        log_str = comm_idx_logs[log_idx]
        comm_rank = log_str[0]
        world_rank = log_str[1]
        match_info = ring_pattern.match(log_str[2])
        if match_info is None:
            continue
        mg = match_info.groups()
        _, _, _, _, ch, prev_node, current_node, next_node = mg
        ch, prev_node, current_node, next_node =\
            (int(ch), int(prev_node), int(current_node), int(next_node))
        assert world_rank == comm_rank_and_idx_to_world_rank(comm_rank, comm_idx)
        if ch >= len(g_rings):
            print('%s, ch=%d, len_rings=%d' % (ring_name, ch, len(g_rings)))
            assert ch < len(g_rings)
        g_rings[ch].add_edge(get_node_name_by_comm_rank_and_idx(prev_node, comm_idx),
                             get_node_name_by_comm_rank_and_idx(current_node, comm_idx))


def process_one_file(full_file_name):
    print('processing node_count=%d, tps=%d, nic_count=%d' % (node_count, tps, nic_count))
    comm_count = tps
    comm_size = node_count * args.gpus_per_node / tps
    first_comm_ranks = comm_first_ranks()
    logs_by_rank = read_file(full_file_name)
    comm_logs = split_log_by_comm(logs_by_rank)
    for idx in range(tps):
        ch_count = nic_count * 2 if node_count > 1 else 32
        all_ring_dict['ring_%dnodes_N%d_T%d_t%d' % (node_count, nic_count, tps, idx)] \
            = [nx.DiGraph() for _ in range(ch_count)]
        all_tree_dict['tree_%dnodes_N%d_T%d_t%d' % (node_count, nic_count, tps, idx)] \
            = [nx.DiGraph() for _ in range(ch_count)]
        for ch in range(ch_count):
            g_ring = all_ring_dict['ring_%dnodes_N%d_T%d_t%d' % (node_count, nic_count, tps, idx)][ch]
            g_tree = all_tree_dict['tree_%dnodes_N%d_T%d_t%d' % (node_count, nic_count, tps, idx)][ch]
            for r in range(8 * node_count // tps):
                node_name = get_node_name_by_comm_rank_and_idx(r, idx)
                g_ring.add_node(node_name)
                g_tree.add_node(node_name)
        extract_test_results(comm_logs[idx])
        extract_tree_info(comm_logs[idx], idx)
        extract_ring_info(comm_logs[idx], idx)


table_header = ['Data Type', 'Data Size', 'Tree LL', 'Tree LL128', 'Tree Simple', 'Ring LL', 'Ring LL128', 'Ring Simple', 'Final Algo', 'Final Proto', 'Final Predict Time', 'Pattern', 'nstepsPerLoop', 'AlgoBW', 'BusBW', 'Time(us)']
comm_space = 3
start_col = 'B'
global col_names
pattern_table = ['Ring', 'RingTwice', 'PipelineFrom', 'PipelineTo', 'TreeUp', 'TreeDown', 'TreeUpDown', 'CollTreeUpDown']


def fill_ws(wb):
    global node_count, nic_count, tps
    node_processed_data = processed_data[node_count]
    comm_idx_data = [None] * tps
    comm_count = 0
    #print('len node_data=%d' % (len(node_processed_data)))
    for data_record in node_processed_data:
        #print('data_record, tps=%d, nic=%d, node_count=%d, len=%d' % (data_record[0].tps, data_record[0].nic, data_record[0].node_count, len(data_record)))
        if data_record[0].tps == tps and data_record[0].nic == nic_count and data_record[0].node_count == node_count:
            assert comm_idx_data[data_record[0].comm_rank] is None
            comm_idx_data[data_record[0].comm_rank] = data_record
            comm_count += 1
    assert comm_count == 0 or comm_count == tps
    if comm_count == 0:
        return
    record_count = len(comm_idx_data[0])
    for i in range(tps):
        assert record_count == len(comm_idx_data[i])
    ws = wb.create_sheet('%dnodes_N%d_T%d' % (node_count, nic_count, tps))
    ws.title = '%dnodes_N%d_T%d' % (node_count, nic_count, tps)
    global col_names
    for i in range(tps):
        start_row = comm_space + (record_count + 1 + comm_space) * i
        tps_data = []
        for field_idx in range(len(table_header)):
            ws[col_names[field_idx] + str(start_row)] = table_header[field_idx]
        for record_idx in range(record_count):
            cid = comm_idx_data[i][record_idx]
            record_fields = ['half', cid.data_size]
            assert len(cid.preds) == 6
            for p in range(6):
                q = 0
                while q < 6:
                    combined = cid.preds[q][0] + ' ' + cid.preds[q][1]
                    if combined == table_header[2 + p]:
                        record_fields.append(cid.preds[q][2])
                        break
                    q += 1
                if q == 6:
                    print('node_count=%d, nic=%d, tps=%d, data_size=%d, %s not found, preds=%s' % (node_count, nic_count, tps, cid.data_size, table_header[2 + p], cid.preds))
                    assert q < 6
            record_fields += [cid.final_algo, cid.final_proto, cid.final_pred_time, pattern_table[cid.pattern], cid.nstepsPerLoop, cid.algo_bw, cid.bus_bw, cid.real_time_us]
            for field_idx in range(len(table_header)):
                ws[col_names[field_idx] + str(start_row + record_idx + 1)] = record_fields[field_idx]
            tps_data.append(record_fields)
        all_data_dict['data_%dnodes_N%d_T%d_t%d' % (node_count, nic_count, tps, i)] = tps_data


def output_xlsx():
    wb = openpyxl.Workbook()
    wb.remove_sheet(wb.active)
    global node_count, nic_count, tps
    for node_count in node_count_array:
        for nic_count in [1, 2, 4, 8]:
            for tps in [1, 2, 4, 8]:
                if node_count == 1 and tps == 8:
                    continue
                fill_ws(wb)
    wb.save(args.output)


def output_rings():
    for name, graph_list in all_ring_dict.items():
        ch_count = len(graph_list)
        for i in range(ch_count):
            g = graph_list[i]
            if g.number_of_edges() == 0:
                continue
            pos = nx.drawing.nx_pydot.graphviz_layout(g, prog="circo")
            new_pos = {}
            plt.figure(figsize=(64, 64))
            for node, (x, y) in pos.items():
                new_pos[node] = (x * 10, y * 10)
            nx.draw(g, pos=new_pos, node_color='yellow', with_labels=False, node_size=1000)
            for node, (x, y) in new_pos.items():
                plt.text(x, y, node, fontsize=20, ha='center', va='center')
            #nx.draw(g, pos)
            fig_name = name + '_ch%d.eps' % (i, )
            plt.savefig(fig_name, format='eps')
            plt.clf()


def output_trees():
    for name, graph_list in all_tree_dict.items():
        ch_count = len(graph_list)
        for i in range(ch_count):
            g = graph_list[i]
            if g.number_of_edges() == 0:
                continue
            pos = nx.drawing.nx_pydot.graphviz_layout(g, prog="dot")
            new_pos = {}
            plt.figure(figsize=(64, 64))
            for node, (x, y) in pos.items():
                new_pos[node] = (x * 10, y * 10)
            nx.draw(g, new_pos, with_labels=False, node_size=5000)
            for node, (x, y) in new_pos.items():
                plt.text(x, y, node, fontsize=80, ha='center', va='center')
            #nx.draw(g, pos)
            fig_name = name + '_ch%d.eps' % (i, )
            plt.savefig(fig_name, format='eps')
            plt.clf()


if __name__ == "__main__":
    args = parse_arguments()
    create_tree_and_ring_patterns()
    processed_data = {}
    all_node_count = {}
    all_data_dict = {'table_header': table_header}
    all_ring_dict = {}
    all_tree_dict = {}
    col_names = [str(chr(ord(start_col[0])+i)) for i in range(len(table_header))]
    g = os.walk(args.input_dir)
    pattern = re.compile(r"nccl_test_(\d+)_(\d+)nodes_T(\d+)_N(\d+)\.log")
    file_count = 0
    for path, dir_list, file_list in g:
        for file_name in file_list:
            if file_count > 1000:
                break
            #print(file_name)
            match = pattern.match(file_name)
            if match is None:
                continue
            full_file_name = os.path.join(path, file_name)
            job_id_str, node_count_str, tps_str, nic_count_str = match.groups()
            job_id, node_count, tps, nic_count = int(job_id_str), int(node_count_str), int(tps_str), int(nic_count_str)
            if node_count != 80 or nic_count != 1 or tps != 1:
                continue
            process_one_file(full_file_name)
            file_count += 1
    # 4 nic_count * 4 tps * 11 data_size for others.
    # 4 nic_count * 3 tps * 11 data_size for 1 node (tps=8 with one node don't need AllReduce).
    node_count_array = [k for k in all_node_count.keys()]
    node_count_array.sort()
    output_xlsx()
    output_rings()
    output_trees()
    #pickle.dump(all_data_dict, open(args.pkl_save_file, "w"))

